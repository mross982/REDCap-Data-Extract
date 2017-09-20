import logging
import sys
import datetime
import export_records as exp
import json
from openpyxl import Workbook
import os.path
from pprint import pprint


def setDates():
    """
This program extracts BHA SUD patients from RedCap who had an encounter the month prior to the date you run this program. Then returns a xlsx file with the same data provided by the project team each month.
    """
    logger.info('***************  Program Start  *******************')

    # Finds the first and last day of the previous month which is used to filter patient encounters and to name the final xlsx file.
    today = datetime.date.today()
    first = today.replace(day=1)
    lastmonth = first - datetime.timedelta(days = 1)
    firstlastmonth = lastmonth.replace(day=1)
    filename = lastmonth.strftime("%B")+'_BHA_SUD_pt_list.xlsx'
    lastmonth = lastmonth.isoformat()
    firstlastmonth = firstlastmonth.isoformat()

    logger.info('Encounter dates filtered from ' + str(firstlastmonth) + ' to ' + str(lastmonth))

    return(firstlastmonth, lastmonth, filename)

def extractData():
    """
extracts data from redcap into a json file
    """
    # extract all data from the RedCap project.
    json_data = exp.getRecords()

    logger.info('Data successfully extracted from RedCap')

    return json_data


def assembleRecords(startDate, stopDate, json_data):

    # Create the initial list of records (dictionaries) with all patients who had an encounter date that is between the first and last day of the previous month.
    initList = []
    for record in json_data:
        for key, value in record.items():
            if key == 'consult_date':
                if (value >= startDate) & (value <= stopDate):
                    newrecord = {
                    'participant_id': record['participant_id'],
                    'last_name': record['last_name'],
                    'first_name': record['first_name'],
                    'dob': record['dob'],
                    'site': record['site'],
                    'consult_date': record['consult_date'],
                    'ref_sud': record['ref_sud'],
                    'ref_bha': '1',
                    'mrn': record['mrn'],
                    'fin': record['fin']
                    }
                    initList.append(newrecord)

    return initList


def correction(initList, json_data):
    """
From the init List, creates a list of unique identifiers from records that need correction. From the json_data, creates a second list that combines the
unique identifiers just discussed with the additional details needed for the final list (i.e. Fin, Mrn, last name, first name). Finally, it replaces the
missing values in the init list with those found in other records of the database.
    """

    # Find and create a list of unique identifiers (participant_id) of all of the entries with blank demographic fields [lname, fname, dob & mrn]
    corrections = []
    for record in initList:
        if record['last_name'] == '':
            corrections.append(record['participant_id'])

    # Compare the list of incomplete demographic records with the entire RedCap database json file and create a new list of records that includes the missing demographic data (corr_dicts).
    cor_dicts = []
    for record in json_data:
        for value in record.values():
            if value in corrections:
                if record['last_name'] != '':
                    cor_record = {
                    'participant_id': record['participant_id'],
                    'last_name': record['last_name'],
                    'first_name': record['first_name'],
                    'dob': record['dob'],
                    'mrn': record['mrn']
                    }
                    cor_dicts.append(cor_record)

    # Compare the two lists of dictionaries (valrecords & cor_dicts) then replace the missing demographic data with that found in the json file.
    for record in initList:
        for k, v in record.items():
            for dic in cor_dicts:
                for k, v in dic.items():
                    if dic['participant_id'] == record['participant_id']:
                        record['last_name'] = dic['last_name']
                        record['first_name'] = dic['first_name']
                        record['dob'] = dic['dob']
                        record['mrn'] = dic['mrn']

    logger.info('Corrected '+str(len(corrections))+' records without patient demographics.')
    logger.info('Found ' + str(len(initList)) + ' total records.')

    return initList


def writedata(finalList, fileName):

    xlorder = {'participant_id': 1, 'last_name': 2, 'first_name': 3, 'dob': 4, 'site': 5, 'consult_date': 6, 'ref_sud': 7,'ref_bha': 8, 'mrn': 9, 'fin': 10}

    wb = Workbook()
    xlsheet = wb.active
    firstrow = 2

    # First row of the output file is the keys to the order dictionary
    for k, v in xlorder.items():
        xlsheet.cell(row = 1, column= xlorder[k]).value = k

    # Places each record into the output file and converts the coded site data into the appropriate site description.
    for record in finalList:
        for k, v in record.items():
            if k in xlorder:
                if k == 'site':
                    if v == '2':
                        v = 'UMCB Floor'
                    else:
                        v = 'UMCB ED'
                xlsheet.cell(row = firstrow, column = xlorder[k]).value = v
        firstrow += 1

    # saves the output file in the data directory
    path = 'data'
    datahome = os.path.join(path, fileName)
    wb.save(datahome)
    logging.info('Data was successfully extracted & reported')


if __name__ == '__main__':

    #define & start the logging process
    logging.basicConfig(format='%(levelname)s:%(asctime)s:%(message)s', datefmt='%m/%d/%Y %I:%M:%S %p', filename='Info.log', level=logging.INFO)
    logger = logging.getLogger(__name__)

    #set custom dates
    startDate = datetime.date(2017, 7, 1)
    startDate = startDate.isoformat()
    stopDate = datetime.date(2017, 8, 1)
    stopDate = stopDate.isoformat()
    fileName = 'July_Data_Verification'

    # OR previous month

    # startDate, stopDate, fileName = setDates()

    json_data = extractData()

    initList = assembleRecords(startDate, stopDate, json_data)

    finalList = correction(initList, json_data)

    fileName = fileName + '.xlsx'
    writedata(finalList, fileName)
